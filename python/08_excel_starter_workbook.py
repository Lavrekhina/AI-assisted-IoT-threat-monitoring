#!/usr/bin/env python3
"""
Build a starter xlsx for the course. One sheet of rows for pivots. KPI sheet if
you already ran steps 03 and 06. Short how to sheet. Needs openpyxl via pip.
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path

import pandas as pd

from features_io import PROJECT_ROOT, read_features_dataframe, resolve_features_path

try:
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError as e:  # pragma: no cover
    raise SystemExit("Install openpyxl first using pip") from e


PIVOT_COLUMNS = [
    "timestamp_utc",
    "ts",
    "device_id",
    "device_type",
    "building",
    "floor",
    "room",
    "firmware_version",
    "protocol",
    "port",
    "event_type",
    "bytes_in",
    "bytes_out",
    "packets",
    "anomaly_score",
    "y_compromise",
    "ground_truth_compromise",
    "hour",
    "day_of_week",
    "dow_name",
    "compromise_type",
]


def _ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)


def _strip_timezone_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    """openpyxl cannot write tz aware datetimes. Store UTC wall time with no tz."""
    out = df.copy()
    for col in out.columns:
        s = out[col]
        if not isinstance(s.dtype, pd.DatetimeTZDtype):
            continue
        out[col] = s.dt.tz_convert("UTC").map(
            lambda t: t.replace(tzinfo=None) if pd.notna(t) else t
        )
    return out


def load_kpi_rows(
    ai_eval_dir: Path, op_dir: Path
) -> list[tuple[str, str]]:
    rows: list[tuple[str, str]] = []
    mj = ai_eval_dir / "metrics.json"
    if mj.exists():
        with mj.open() as f:
            m = json.load(f)
        for k, v in m.items():
            rows.append((k, str(v)))
    else:
        rows.append(
            (
                "note",
                "Run 03_ai_evaluation.py first to get metrics.json then run this script again",
            )
        )
    kpi = op_dir / "kpi_delay_summary.csv"
    if kpi.exists():
        d = pd.read_csv(kpi)
        for _, r in d.iterrows():
            rows.append((str(r.get("metric", "")), str(r.get("value", ""))))
    return rows


def main() -> None:
    ap = argparse.ArgumentParser(description="Export starter Excel workbook for pivots and KPIs")
    ap.add_argument(
        "--features",
        default="artifacts/features.parquet",
        help="Parquet or CSV from 01 (CSV if parquet missing)",
    )
    ap.add_argument("--out", default="artifacts/excel/iot_telemetry_pivot_starter.xlsx")
    ap.add_argument("--max-rows", type=int, default=0, help="zero means all rows")
    args = ap.parse_args()

    out_path = Path(args.out)
    _ensure_dir(out_path.parent)

    fp = resolve_features_path(args.features)
    print(f"Using features: {fp}")
    df = read_features_dataframe(fp)
    use = [c for c in PIVOT_COLUMNS if c in df.columns]
    df = df[use]
    if args.max_rows and len(df) > args.max_rows:
        df = df.head(args.max_rows)
    df = _strip_timezone_for_excel(df)

    wb = Workbook()
    ws = wb.active
    if ws is not None:
        ws.title = "data_for_pivot"
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

    whelp = wb.create_sheet("pivot_howto")
    lines = [
        "Pick a cell on data_for_pivot. Use Insert then Pivot table on desktop Excel",
        "Put device_type building floor room day_of_week hour in rows or columns as you like",
        "Add count of device_id sum of y_compromise median anomaly score sum of bytes out",
        "Add slicers for hour device_type compromise_type. Add a helper column for top decile if you want",
        "KPI cards on another sheet. Point cells at the pivot or at K_summary",
        "Optional. Add a flag when anomaly score is above a cell T1 and compare to y for false pos neg study",
    ]
    for i, line in enumerate(lines, start=1):
        whelp.cell(row=i, column=1, value=line)

    wk = wb.create_sheet("K_summary")
    wk.cell(row=1, column=1, value="metric")
    wk.cell(row=1, column=2, value="value")
    kpi = load_kpi_rows(PROJECT_ROOT / "artifacts" / "ai_eval", PROJECT_ROOT / "artifacts" / "operational")
    for i, (a, b) in enumerate(kpi, start=2):
        wk.cell(row=i, column=1, value=a)
        wk.cell(row=i, column=2, value=b)

    wb.save(out_path)
    print(f"Wrote {out_path} ({len(df)} data rows, {len(use)} columns)")


if __name__ == "__main__":
    main()
